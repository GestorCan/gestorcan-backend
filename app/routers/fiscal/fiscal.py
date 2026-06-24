from fastapi import APIRouter, Request, Depends,Form
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.factura import Factura
from app.models.verifactu_registro import VeriFactuRegistro
from app.verifactu.constants import (ESTADO_FIRMADO,ESTADO_ACEPTADO,ESTADO_ERROR)
from fastapi.templating import Jinja2Templates
from datetime import date
from app.models.gasto import Gasto
from pathlib import Path
from openpyxl import Workbook
from fastapi.responses import RedirectResponse,FileResponse
from app.models.proveedor import Proveedor
from shutil import copy2
import zipfile
from reportlab.platypus import (SimpleDocTemplate,Paragraph,Spacer)
from app.models.residencias import Residencia
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from datetime import date
from app.models.factura import Factura
from app.models.gasto import Gasto
import os







router = APIRouter()
templates = Jinja2Templates(directory="app/templates")



INFORMES_FISCAL_DIR = (
    Path("static")
    / "informes"
    / "fiscal"
    / "iva"
)




def generar_resumen_fiscal_pdf(
    ruta_pdf,
    nombre_residencia,
    cif_residencia,
    logo_path,
    ejercicio,
    trimestre,
    fecha_generacion,
    facturas_emitidas,
    gastos,
    base_emitidas,
    iva_emitidas,
    total_emitidas,
    base_recibidas,
    iva_recibidas,
    total_recibidas,
        total_retenciones=0,
        num_facturas_retencion=0,
        base_retenciones=0,
        iva_neto=0,

):
    c = canvas.Canvas(str(ruta_pdf), pagesize=landscape(A4))
    width, height = landscape(A4)

    azul = colors.HexColor("#0f3a75")
    gris = colors.HexColor("#6b7280")
    verde = colors.HexColor("#3b8d2a")
    naranja = colors.HexColor("#f97316")
    morado = colors.HexColor("#7c3aed")

    # Fondo
    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, fill=True, stroke=False)

    from pathlib import Path


    # Logo
    if logo_path and Path(logo_path).exists():
        c.drawImage(
            str(logo_path),
            width / 2 - 55,
            height - 85,
            width=110,
            height=45,
            preserveAspectRatio=True,
            mask="auto"
        )

    y_nombre = height - 115
    y_cif = height - 132
    y_caja = height - 225






    # Cabecera
    c.setFillColor(azul)
    c.setFont("Helvetica-Bold", 28)
    c.drawCentredString(width / 2, y_nombre, nombre_residencia.upper())

    if cif_residencia:
        c.setFont("Helvetica", 11)
        c.drawCentredString(width / 2, y_cif, f"CIF: {cif_residencia}")

    # Caja título
    c.setStrokeColor(azul)
    c.setLineWidth(1.5)
    c.roundRect(25, y_caja, width - 50, 80, 12, stroke=True, fill=False)

    c.setFont("Helvetica-Bold", 22)
    c.setFillColor(azul)
    c.drawCentredString(width / 2, y_caja + 52, "PAQUETE ASESOR TRIMESTRAL")

    c.setFont("Helvetica-Bold", 15)
    c.drawCentredString(width / 2, y_caja + 28, f"{trimestre}º Trimestre {ejercicio}")

    c.setFont("Helvetica", 10)
    c.setFillColor(colors.black)
    c.drawCentredString(width / 2, y_caja + 10, f"Generado el {fecha_generacion}")

    # Título sección
    y = height - 210
    c.setFillColor(azul)
    c.setFont("Helvetica-Bold", 15)
    c.drawString(35, y, "")
    #c.line(190, y + 4, width - 35, y + 4)

    # Tarjetas superiores
    def tarjeta(x, y, w, h, titulo, color, lineas, total_label, total_valor):
        c.setStrokeColor(color)
        c.setFillColor(colors.white)
        c.roundRect(x, y, w, h, 10, stroke=True, fill=True)

        c.setFillColor(color)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x + 18, y + h - 28, titulo)

        c.setFillColor(colors.black)
        c.setFont("Helvetica", 10)

        yy = y + h - 42
        for label, valor in lineas:
            c.drawString(x + 18, yy, label)
            c.drawRightString(x + w - 18, yy, valor)
            yy -= 17

        c.setStrokeColor(color)
        c.line(x + 18, y + 35, x + w - 18, y + 35)


        c.setFillColor(color)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x + 18, y + 18, total_label)
        c.drawRightString(x + w - 18, y + 18, total_valor)

    tarjeta(
        35, 230, 250, 115,
        "FACTURAS EMITIDAS",
        verde,
        [
            ("Nº facturas", str(len(facturas_emitidas))),
            ("Base imponible", f"{base_emitidas:.2f} €"),
            ("IVA repercutido", f"{iva_emitidas:.2f} €"),
        ],
        "Total facturado",
        f"{total_emitidas:.2f} €"
    )

    tarjeta(
        305, 230, 250, 115,
        "FACTURAS RECIBIDAS",
        naranja,
        [
            ("Nº facturas", str(len(gastos))),
            ("Base imponible", f"{base_recibidas:.2f} €"),
            ("IVA soportado", f"{iva_recibidas:.2f} €"),
        ],
        "Total gastos",
        f"{total_recibidas:.2f} €"
    )

    tarjeta(
        575, 230, 230, 115,
        "RETENCIONES",
        morado,
        [
            ("Facturas", str(num_facturas_retencion)),
            ("Base sujeta", f"{base_retenciones:.2f} €"),
        ],
        "Total retenciones",
        f"{total_retenciones:.2f} €"
    )

    # Resultado IVA
    resultado_x = 35
    resultado_y = 90
    resultado_w = 520
    resultado_h = 115

    c.setFillColor(colors.HexColor("#eff6ff"))
    c.setStrokeColor(azul)
    c.roundRect(resultado_x, resultado_y, resultado_w, resultado_h, 10, fill=True, stroke=True)

    c.setFillColor(azul)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(resultado_x + 20, resultado_y + resultado_h - 25, "RESULTADO FISCAL DEL PERIODO")

    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)

    c.drawString(resultado_x + 40, resultado_y + 65, "IVA repercutido")
    c.drawRightString(resultado_x + resultado_w - 40, resultado_y + 65, f"{iva_emitidas:.2f} €")

    c.drawString(resultado_x + 40, resultado_y + 45, "IVA soportado")
    c.drawRightString(resultado_x + resultado_w - 40, resultado_y + 45, f"{iva_recibidas:.2f} €")

    c.setStrokeColor(azul)
    c.line(resultado_x + 40, resultado_y + 35, resultado_x + resultado_w - 40, resultado_y + 35)

    estado = "IVA A INGRESAR" if iva_neto >= 0 else "IVA A COMPENSAR"

    c.setFillColor(azul)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(resultado_x + 40, resultado_y + 15, estado)
    c.drawRightString(resultado_x + resultado_w - 40, resultado_y + 15, f"{abs(iva_neto):.2f} €")

    # Documentación incluida
    docs_x = 575
    docs_y = 90
    docs_w = 230
    docs_h = 115

    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor("#cbd5e1"))
    c.roundRect(docs_x, docs_y, docs_w, docs_h, 10, fill=True, stroke=True)

    c.setFillColor(azul)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(docs_x + 18, docs_y + docs_h - 25, "DOCUMENTACIÓN INCLUIDA")

    c.setFillColor(colors.black)
    c.setFont("Helvetica", 8.5)

    docs = [
        "✓ Libro de facturas emitidas",
        "✓ Libro de facturas recibidas",
        "✓ Relación de retenciones",
        "✓ Facturas emitidas en PDF",
        "✓ Justificantes de gastos",
    ]

    yy = docs_y + docs_h - 45

    for d in docs:
        c.drawString(docs_x + 18, yy, d)
        yy -= 14
    # Pie
    c.setFillColor(colors.HexColor("#eff6ff"))
    c.roundRect(35, 35, 560, 40, 8, fill=True, stroke=False)

    c.setFillColor(azul)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(55, 58, "Documento generado automáticamente.")

    c.setFillColor(colors.black)
    c.setFont("Helvetica", 8)
    c.drawString(
        55,
        45,
        "Este resumen acompaña a la documentación fiscal incluida en el paquete asesor correspondiente al periodo indicado."
    )

    c.setFillColor(azul)
    c.setFont("Helvetica-Bold", 14)

    # Logo
    logo_gestorcan = os.path.join(
        "static",
        "logos",
        "logo_1.png"
    )

    if os.path.exists(logo_gestorcan):
        c.drawImage(
            logo_gestorcan,
            width - 140,  # mueve izquierda/derecha
            50 , # misma altura que el texto
            width=22,
            height=22,
            preserveAspectRatio=True,
            mask='auto'
        )

    # Texto
    c.drawRightString(width - 45, 55, "GestorCan")

    c.save()



@router.post("/paquete-asesor/crear")
def crear_paquete_asesor(
    ejercicio: int = Form(...),
    trimestre: int = Form(...),
    db: Session = Depends(get_db)
):

    fecha_desde, fecha_hasta = rango_trimestre(
        ejercicio,
        trimestre
    )

    nombre_carpeta = f"{ejercicio}_T{trimestre}"
    logo_residencia = "static/logos/residencia_1.png"

    carpeta_trimestre = (
            INFORMES_FISCAL_DIR
            / nombre_carpeta
    )

    carpeta_excel = (
            carpeta_trimestre
            / "Excel"
    )

    carpeta_emitidas = (
            carpeta_trimestre
            / "Facturas Emitidas PDF"
    )

    carpeta_recibidas = (
            carpeta_trimestre
            / "Facturas Recibidas"
    )

    carpeta_excel.mkdir(
        parents=True,
        exist_ok=True
    )

    carpeta_emitidas.mkdir(
        parents=True,
        exist_ok=True
    )

    carpeta_recibidas.mkdir(
        parents=True,
        exist_ok=True
    )

    # ------------------------
    # FACTURAS EMITIDAS
    # ------------------------

    facturas_emitidas = (
        db.query(Factura)
        .filter(Factura.fecha >= fecha_desde)
        .filter(Factura.fecha <= fecha_hasta)
        .all()
    )

    wb = Workbook()
    ws = wb.active

    ws.title = "Facturas Emitidas"

    ws.append([
        "Fecha",
        "Número factura",
        "Tipo",
        "Nombre / Apellidos",
        "CIF",
        "Base Imponible",
        "% IVA",
        "Importe IVA",
        "Total Factura"
    ])

    for factura in facturas_emitidas:
        cliente = factura.cliente

        nombre_cliente = ""
        cif_cliente = ""

        if cliente:
            nombre_cliente = (
                f"{cliente.nombre or ''} "
                f"{cliente.apellidos or ''}"
            ).strip()

            cif_cliente = cliente.dni or ""

        base = float(factura.base_imponible or 0)
        iva = float(factura.iva or 0)
        total = float(factura.total or 0)

        porcentaje_iva = 0

        if base > 0:
            porcentaje_iva = round((iva / base) * 100, 2)

        ws.append([
            factura.fecha.strftime("%d/%m/%Y"),
            factura.numero_factura,
            factura.tipo_factura or "ordinaria",
            nombre_cliente,
            cif_cliente,
            base,
            porcentaje_iva,
            iva,
            total
        ])

    archivo_emitidas = (
        carpeta_excel
        / f"facturas_emitidas_{ejercicio}_T{trimestre}.xlsx"
    )

    wb.save(archivo_emitidas)

    # ------------------------
    # FACTURAS RECIBIDAS
    # ------------------------

    gastos = (
        db.query(Gasto)
        .filter(Gasto.tipo_documento == "factura")
        .filter(Gasto.fecha >= fecha_desde)
        .filter(Gasto.fecha <= fecha_hasta)
        .filter(Gasto.incluida_paquete_asesor == False)
        .all()
    )

    wb = Workbook()
    ws = wb.active

    ws.title = "Facturas Recibidas"



    cif_proveedor = ""


    ws.append([
        "Fecha",
        "Número factura",
        "Proveedor",
        "CIF",
        "Tipo gasto",
        "Concepto",
        "Base imponible",
        "% IVA",
        "Importe IVA",
        "% Retención",
        "Importe retención",
        "Total factura"
    ])

    for gasto in gastos:

        proveedor_obj = None
        cif_proveedor = ""

        if gasto.proveedor_id:
            proveedor_obj = (
                db.query(Proveedor)
                .filter(Proveedor.id == gasto.proveedor_id)
                .first()
            )

        if not proveedor_obj and gasto.proveedor:
            proveedor_obj = (
                db.query(Proveedor)
                .filter(Proveedor.nombre == gasto.proveedor)
                .first()
            )

        if proveedor_obj:
            cif_proveedor = proveedor_obj.cif or ""

        ws.append([
            gasto.fecha.strftime("%d/%m/%Y"),
            gasto.numero_factura or "",
            gasto.proveedor,
            cif_proveedor,
            gasto.tipo_gasto,
            gasto.concepto or "",
            float(gasto.base_imponible or 0),
            float(gasto.porcentaje_iva or 0),
            float(gasto.importe_iva or 0),
            float(gasto.porcentaje_retencion or 0),
            float(gasto.importe_retencion or 0),
            float(gasto.total_factura or 0)
        ])

    archivo_recibidas = (
        carpeta_excel
        / f"facturas_recibidas_{ejercicio}_T{trimestre}.xlsx"
    )

    wb.save(archivo_recibidas)

    # ------------------------
    # RETENCIONES
    # ------------------------

    wb = Workbook()
    ws = wb.active

    ws.title = "Retenciones"

    ws.append([
        "Fecha",
        "Proveedor",
        "CIF",
        "Concepto",
        "Retencion"
    ])

    for gasto in gastos:

        if (gasto.importe_retencion or 0) > 0:

            ws.append([
                gasto.fecha.strftime("%d/%m/%Y"),
                gasto.proveedor,
                cif_proveedor,
                gasto.concepto,
                float(gasto.importe_retencion)
            ])

    archivo_retenciones = (
        carpeta_excel
        / f"retenciones_{ejercicio}_T{trimestre}.xlsx"
    )

    wb.save(archivo_retenciones)

    # ==================================
    # TOTALES FISCALES
    # ==================================

    base_emitidas = sum(
        float(f.base_imponible or 0)
        for f in facturas_emitidas
    )

    iva_emitidas = sum(
        float(f.iva or 0)
        for f in facturas_emitidas
    )

    total_emitidas = sum(
        float(f.total or 0)
        for f in facturas_emitidas
    )

    base_recibidas = sum(
        float(g.base_imponible or 0)
        for g in gastos
    )

    iva_recibidas = sum(
        float(g.importe_iva or 0)
        for g in gastos
    )

    total_recibidas = sum(
        float(g.total_factura or 0)
        for g in gastos
    )

    total_retenciones = sum(
        float(g.importe_retencion or 0)
        for g in gastos
    )

    iva_neto = iva_emitidas - iva_recibidas

    # ==================================
    # RESUMEN FISCAL
    # ==================================

    archivo_resumen = (
            carpeta_trimestre
            / f"Resumen_Fiscal_{ejercicio}_T{trimestre}.pdf"
    )

    residencia = db.query(Residencia).first()

    nombre_residencia = (
        residencia.nombre
        if residencia and residencia.nombre
        else "GestorCan"
    )

    cif_residencia = (
        residencia.cif
        if residencia and hasattr(residencia, "cif")
        else ""
    )
    total_retenciones = sum(
        float(g.importe_retencion or 0)
        for g in gastos
    )

    num_facturas_retencion = sum(
        1
        for g in gastos
        if float(g.importe_retencion or 0) > 0
    )

    base_retenciones = sum(
        float(g.base_imponible or 0)
        for g in gastos

        if float(g.importe_retencion or 0) > 0
    )

    iva_neto = iva_emitidas - iva_recibidas


    generar_resumen_fiscal_pdf(
        archivo_resumen,
        nombre_residencia,
        cif_residencia,
        logo_residencia,
        ejercicio,
        trimestre,
        date.today().strftime("%d/%m/%Y"),
        facturas_emitidas,
        gastos,
        base_emitidas,
        iva_emitidas,
        total_emitidas,
        base_recibidas,
        iva_recibidas,
        total_recibidas,
        total_retenciones=total_retenciones,
        num_facturas_retencion=num_facturas_retencion,
        base_retenciones=base_retenciones,
        iva_neto=iva_neto,
    )

    print("PDF RESUMEN CREADO:", archivo_resumen)





    # ==================================
    # COPIAR PDF FACTURAS EMITIDAS
    # ==================================

    pdfs_copiados = 0

    for factura in facturas_emitidas:

        if not factura.pdf_path:
            continue

        ruta_pdf = Path(factura.pdf_path)

        # si guarda rutas tipo /static/...
        if not ruta_pdf.exists():
            ruta_pdf = Path(
                factura.pdf_path.replace(
                    "/static/",
                    "static/"
                )
            )

        if ruta_pdf.exists():
            copy2(
                ruta_pdf,
                carpeta_emitidas / ruta_pdf.name
            )

            pdfs_copiados += 1


        # ==================================
        # COPIAR ADJUNTOS GASTOS
        # ==================================

        adjuntos_copiados = 0

        adjuntos_copiados = 0

        for gasto in gastos:

            if not gasto.archivo_path:
                continue

            ruta_archivo = Path(gasto.archivo_path)

            if not ruta_archivo.exists():
                ruta_archivo = Path(
                    gasto.archivo_path.replace(
                        "/static/",
                        "static/"
                    )
                )

            if ruta_archivo.exists():
                copy2(
                    ruta_archivo,
                    carpeta_recibidas / ruta_archivo.name
                )

                adjuntos_copiados += 1
                print("PDFS COPIADOS:", pdfs_copiados)
                print("ADJUNTOS COPIADOS:", adjuntos_copiados)

    # ==================================
    # CREAR ZIP
    # ==================================

    ruta_zip = (
            carpeta_trimestre
            / f"paquete_asesor_{ejercicio}_T{trimestre}.zip"
    )

    with zipfile.ZipFile(
            ruta_zip,
            "w",
            zipfile.ZIP_DEFLATED
    ) as zipf:

        for archivo in carpeta_trimestre.rglob("*"):

            if archivo == ruta_zip:
                continue

            if archivo.is_file():
                zipf.write(
                    archivo,
                    archivo.relative_to(carpeta_trimestre)
                )

    print("ZIP CREADO:", ruta_zip)
    from datetime import datetime

    for gasto in gastos:
        gasto.incluida_paquete_asesor = True
        gasto.ejercicio_paquete_asesor = ejercicio
        gasto.trimestre_paquete_asesor = trimestre
        gasto.fecha_inclusion_paquete = datetime.now()

    db.commit()

    return FileResponse(
        path=ruta_zip,
        filename=ruta_zip.name,
        media_type="application/zip"
    )



def rango_trimestre(ejercicio: int, trimestre: int):
    if trimestre == 1:
        return date(ejercicio, 1, 1), date(ejercicio, 3, 31)

    if trimestre == 2:
        return date(ejercicio, 4, 1), date(ejercicio, 6, 30)

    if trimestre == 3:
        return date(ejercicio, 7, 1), date(ejercicio, 9, 30)

    if trimestre == 4:
        return date(ejercicio, 10, 1), date(ejercicio, 12, 31)

    raise ValueError("Trimestre no válido")

@router.get("/paquete-asesor")
def paquete_asesor(
    request: Request
):
    ejercicio_actual = date.today().year

    return templates.TemplateResponse(
        request=request,
        name="fiscal/paquete_asesor.html",
        context={
            "ejercicio_actual": ejercicio_actual
        }
    )

@router.post("/paquete-asesor/generar")
def generar_paquete_asesor_resumen(
    request: Request,
    ejercicio: int = Form(...),
    trimestre: int = Form(...),
    db: Session = Depends(get_db)
):
    fecha_desde, fecha_hasta = rango_trimestre(
        ejercicio,
        trimestre
    )

    facturas_emitidas = (
        db.query(Factura)
        .filter(Factura.fecha >= fecha_desde)
        .filter(Factura.fecha <= fecha_hasta)
        .all()
    )

    facturas_recibidas = (
        db.query(Gasto)
        .filter(Gasto.tipo_documento == "factura")
        .filter(Gasto.fecha >= fecha_desde)
        .filter(Gasto.fecha <= fecha_hasta)
        .all()
    )

    base_emitidas = sum(float(f.base_imponible or 0) for f in facturas_emitidas)
    iva_emitidas = sum(float(f.iva or 0) for f in facturas_emitidas)
    total_emitidas = sum(float(f.total or 0) for f in facturas_emitidas)

    base_recibidas = sum(float(g.base_imponible or 0) for g in facturas_recibidas)
    iva_recibidas = sum(float(g.importe_iva or 0) for g in facturas_recibidas)
    total_recibidas = sum(float(g.total_factura or 0) for g in facturas_recibidas)

    total_retenciones = sum(float(g.importe_retencion or 0) for g in facturas_recibidas)
    num_facturas_retencion = sum(1 for g in facturas_recibidas if float(g.importe_retencion or 0) > 0 )

    base_retenciones = sum(float(g.base_imponible or 0)for g in facturas_recibidas if float(g.importe_retencion or 0) > 0    )

    iva_neto = iva_emitidas - iva_recibidas

    return templates.TemplateResponse(
        request=request,
        name="fiscal/paquete_asesor_resumen.html",
        context={
            "ejercicio": ejercicio,
            "trimestre": trimestre,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,

            "facturas_emitidas": facturas_emitidas,
            "facturas_recibidas": facturas_recibidas,

            "num_emitidas": len(facturas_emitidas),
            "base_emitidas": base_emitidas,
            "iva_emitidas": iva_emitidas,
            "total_emitidas": total_emitidas,

            "num_recibidas": len(facturas_recibidas),
            "base_recibidas": base_recibidas,
            "iva_recibidas": iva_recibidas,
            "total_recibidas": total_recibidas,

            "total_retenciones": total_retenciones,
            "iva_neto": iva_neto,
        }
    )






@router.get("/liquidaciones")





def liquidaciones_iva(
    request: Request,
    db: Session = Depends(get_db)
):

    hoy = date.today()

    if hoy.month <= 3:
        trimestre = "1º Trimestre"
        fecha_inicio = date(hoy.year, 1, 1)
        fecha_fin = date(hoy.year, 3, 31)

    elif hoy.month <= 6:
        trimestre = "2º Trimestre"
        fecha_inicio = date(hoy.year, 4, 1)
        fecha_fin = date(hoy.year, 6, 30)

    elif hoy.month <= 9:
        trimestre = "3º Trimestre"
        fecha_inicio = date(hoy.year, 7, 1)
        fecha_fin = date(hoy.year, 9, 30)

    else:
        trimestre = "4º Trimestre"
        fecha_inicio = date(hoy.year, 10, 1)
        fecha_fin = date(hoy.year, 12, 31)

    facturas = (
        db.query(Factura)
        .filter(
            Factura.fecha >= fecha_inicio,
            Factura.fecha <= fecha_fin
        )
        .all()
    )

    base_total = sum(float(f.base_imponible or 0) for f in facturas)
    iva_repercutido = sum(float(f.iva or 0) for f in facturas)

    gastos = (
        db.query(Gasto)
        .filter(
            Gasto.fecha >= fecha_inicio,
            Gasto.fecha <= fecha_fin
        )
        .all()
    )

    base_gastos = sum(float(g.base_imponible or 0) for g in gastos)
    iva_soportado = sum(float(g.importe_iva or 0) for g in gastos)

    resultado = iva_repercutido - iva_soportado

    if resultado >= 0:
        estado_resultado = "A ingresar"
        importe_resultado = resultado
    else:
        estado_resultado = "A compensar"
        importe_resultado = abs(resultado)

    def formato_es(valor):
        return (
            "{:,.2f}".format(valor)
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )

    base_total_fmt = formato_es(base_total)
    iva_repercutido_fmt = formato_es(iva_repercutido)
    base_gastos_fmt = formato_es(base_gastos)
    iva_soportado_fmt = formato_es(iva_soportado)
    importe_resultado_fmt = formato_es(importe_resultado)


    return templates.TemplateResponse(
        request=request,
        name="fiscal/liquidaciones.html",
        context={
            "base_total": base_total,
            "iva_repercutido": iva_repercutido,
            "base_gastos": base_gastos,
            "iva_soportado": iva_soportado,
            "resultado": resultado,

            "base_total_fmt": base_total_fmt,
            "iva_repercutido_fmt": iva_repercutido_fmt,
            "base_gastos_fmt": base_gastos_fmt,
            "iva_soportado_fmt": iva_soportado_fmt,
            "importe_resultado_fmt": importe_resultado_fmt,

            "importe_resultado": importe_resultado,
            "estado_resultado": estado_resultado,
            "trimestre": trimestre,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "num_facturas": len(facturas),
            "num_gastos": len(gastos),
        }
    )





@router.get("/libros-registro")
def libros_registro(
    request: Request,
    db: Session = Depends(get_db)
):
    facturas = (
        db.query(Factura)
        .order_by(Factura.fecha.desc())
        .all()
    )

    total_base = sum(float(f.base_imponible or 0) for f in facturas)
    total_iva = sum(float(f.iva or 0) for f in facturas)
    total_facturas = sum(float(f.total or 0) for f in facturas)

    return templates.TemplateResponse(
        request=request,
        name="fiscal/libros_registro.html",
        context={
            "facturas": facturas,
            "total_base": total_base,
            "total_iva": total_iva,
            "total_facturas": total_facturas,
        }
    )

@router.get("")
def fiscal_dashboard(
    request: Request,
    db: Session = Depends(get_db)
):
    facturas = (
        db.query(Factura)
        .order_by(Factura.fecha.desc(), Factura.id.desc())
        .all()
    )

    registros = (
        db.query(VeriFactuRegistro)
        .order_by(VeriFactuRegistro.id.desc())
        .all()
    )

    registros_por_factura = {
        r.factura_id: r
        for r in registros
    }

    total_registros = len(registros)

    firmados = sum(
        1 for r in registros
        if r.estado == ESTADO_FIRMADO
    )

    aceptados = sum(
        1 for r in registros
        if r.estado == ESTADO_ACEPTADO
    )

    errores = sum(
        1 for r in registros
        if r.estado == ESTADO_ERROR
    )

    return templates.TemplateResponse(
        request=request,
        name="fiscal/dashboard.html",
        context={
            "facturas": facturas,
            "registros_por_factura": registros_por_factura,

            "total_registros": total_registros,
            "firmados": firmados,
            "aceptados": aceptados,
            "errores": errores
        }
    )